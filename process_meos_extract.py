#!/usr/bin/env python3
import argparse
import math
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def _open_excel(path: Path) -> pd.ExcelFile:
    try:
        return pd.ExcelFile(path)
    except Exception as exc:  # pragma: no cover - fallback for invalid engine detection
        if "io.excel.zip.reader" in str(exc):
            try:
                return pd.ExcelFile(path, engine="openpyxl")
            except Exception as inner:
                raise ValueError(
                    f"{path} appears to be a ZIP archive rather than a supported "
                    "Excel workbook. Ensure the file is a valid .xlsx/.xlsm workbook."
                ) from inner
        raise


def _load_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    excel = _open_excel(path)
    resolved_sheet = sheet_name
    if sheet_name not in excel.sheet_names:
        matching_sheets = []
        for candidate in excel.sheet_names:
            preview = pd.read_excel(excel, sheet_name=candidate, nrows=0)
            if sheet_name in preview.columns:
                matching_sheets.append(candidate)
        if len(matching_sheets) == 1:
            resolved_sheet = matching_sheets[0]
        else:
            available = ", ".join(excel.sheet_names)
            if matching_sheets:
                matches = ", ".join(matching_sheets)
                raise ValueError(
                    f"Worksheet named '{sheet_name}' not found in {path}. "
                    f"Multiple worksheets contain '{sheet_name}': {matches}."
                )
            raise ValueError(
                f"Worksheet named '{sheet_name}' not found in {path}. "
                f"Available worksheets: {available or 'None'}."
            )
    time_col = "time_iso_utc"
    df = pd.read_excel(excel, sheet_name=resolved_sheet)
    if df.empty:
        print(f"Warning: sheet '{sheet_name}' in {path} is empty; skipping data.")
        empty_frame = pd.DataFrame(columns=[time_col, sheet_name])
        return empty_frame
    if time_col not in df.columns:
        raise ValueError(
            f"Sheet '{sheet_name}' in {path} is missing required '{time_col}' column."
        )
    if sheet_name not in df.columns:
        raise ValueError(
            f"Sheet '{resolved_sheet}' in {path} is missing required '{sheet_name}' column."
        )
    data = df[[time_col, sheet_name]].copy()
    data[time_col] = pd.to_datetime(data[time_col], errors="coerce", utc=True)
    data = data.dropna(subset=[time_col]).copy()
    data[time_col] = data[time_col].dt.floor("s")
    data = data.sort_values(time_col)
    data = data.drop_duplicates(subset=[time_col], keep="first")
    return data


def _process_file(path: Path) -> pd.DataFrame:
    snr = _load_sheet(path, "5_10_signal_noise_ratio")
    azimuth = _load_sheet(path, "6_1_azimuth")
    elevation = _load_sheet(path, "6_2_elevation")

    merged = snr.merge(azimuth, on="time_iso_utc", how="inner").merge(
        elevation, on="time_iso_utc", how="inner"
    )
    merged["orbit"] = _extract_orbit(path)
    merged = merged.sort_values("time_iso_utc")
    return _trim_final_high_elevation(merged)


def _trim_final_high_elevation(
    frame: pd.DataFrame, threshold: float = 5.0
) -> pd.DataFrame:
    if frame.empty:
        return frame
    elevation = pd.to_numeric(frame["6_2_elevation"], errors="coerce")
    valid_mask = elevation <= threshold
    if not valid_mask.any():
        return frame.iloc[0:0].copy()
    last_valid_index = valid_mask[valid_mask].index[-1]
    return frame.loc[:last_valid_index].copy()


def _extract_start_time(path: Path) -> pd.Timestamp:
    excel = _open_excel(path)
    if "__meta__" not in excel.sheet_names:
        available = ", ".join(excel.sheet_names)
        raise ValueError(
            f"Worksheet named '__meta__' not found in {path}. "
            f"Available worksheets: {available or 'None'}."
        )
    meta = pd.read_excel(excel, sheet_name="__meta__")
    if meta.empty:
        raise ValueError(f"Sheet '__meta__' in {path} is empty.")
    start_time = None
    if "start_time_utc" in meta.columns:
        values = meta["start_time_utc"].dropna()
        if not values.empty:
            start_time = values.iloc[0]
    if start_time is None:
        lower_cols = [str(col).strip().lower() for col in meta.columns]
        for _, row in meta.iterrows():
            for idx, col in enumerate(meta.columns):
                cell = row[col]
                if isinstance(cell, str) and cell.strip().lower() == "start_time_utc":
                    next_value = None
                    if idx + 1 < len(meta.columns):
                        next_value = row[meta.columns[idx + 1]]
                    elif "value" in lower_cols:
                        value_idx = lower_cols.index("value")
                        next_value = row[meta.columns[value_idx]]
                    start_time = next_value
                    break
            if start_time is not None:
                break
    start_time = pd.to_datetime(start_time, errors="coerce", utc=True)
    if pd.isna(start_time):
        raise ValueError(f"Unable to parse start_time_utc from {path}.")
    return start_time


def _extract_orbit(path: Path) -> int:
    match = re.search(r"orbit[_-]?(\d+)", path.stem, flags=re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Unable to determine orbit from filename '{path.name}'. "
            "Expected pattern like 'orbit_2648'."
        )
    return int(match.group(1))


def _bin_series(series: pd.Series, tolerance: float) -> pd.Series:
    if tolerance <= 0:
        return series
    return (series / tolerance).round() * tolerance


def _max_center_elevation(series: pd.Series, center_fraction: float = 0.5) -> float:
    if isinstance(series, pd.DataFrame):
        series = pd.Series(series.to_numpy().ravel())
    values = pd.to_numeric(series, errors="coerce").dropna().reset_index(drop=True)
    if values.empty:
        return float("nan")
    if not (0 < center_fraction <= 1):
        center_fraction = 0.5
    window = max(1, int(round(len(values) * center_fraction)))
    start = max(0, (len(values) - window) // 2)
    end = start + window
    return float(values.iloc[start:end].max())


def _build_correlation_table(
    frames: List[pd.DataFrame], azimuth_tolerance: float, elevation_tolerance: float
) -> pd.DataFrame:
    aligned_frames: List[pd.DataFrame] = []
    for frame in frames:
        orbit = int(frame["orbit"].iloc[0])
        aligned = frame[
            ["6_1_azimuth", "6_2_elevation", "5_10_signal_noise_ratio"]
        ].copy()
        aligned["azimuth_bin"] = _bin_series(aligned["6_1_azimuth"], azimuth_tolerance)
        aligned["elevation_bin"] = _bin_series(
            aligned["6_2_elevation"], elevation_tolerance
        )
        aligned = (
            aligned.groupby(["azimuth_bin", "elevation_bin"], as_index=False)
            .agg({"5_10_signal_noise_ratio": "mean"})
            .rename(
                columns={
                    "5_10_signal_noise_ratio": f"Orbit {orbit} SNR",
                }
            )
        )
        aligned_frames.append(aligned)

    if not aligned_frames:
        return pd.DataFrame()

    merged = aligned_frames[0]
    for aligned in aligned_frames[1:]:
        merged = merged.merge(aligned, on=["azimuth_bin", "elevation_bin"], how="inner")

    return merged.sort_values(["azimuth_bin", "elevation_bin"]).reset_index(drop=True)


def _add_group_chart_sheet(
    workbook,
    source_sheet,
    sheet_name: str,
    output_frame: pd.DataFrame,
) -> None:
    if output_frame.empty:
        return

    chart_sheet = workbook.create_sheet(title=f"{sheet_name} Chart")
    chart = ScatterChart()
    chart.title = f"{sheet_name} SNR vs Elevation"
    chart.x_axis.title = "Elevation (deg)"
    chart.y_axis.title = "Signal-to-noise ratio"
    chart.x_axis.scaling.min = 5
    elevation_cols = [
        col for col in output_frame.columns if col.endswith("6_2_elevation")
    ]
    if elevation_cols:
        max_elevation = max(
            (
                _max_center_elevation(output_frame[col])
                for col in elevation_cols
            ),
            default=float("nan"),
        )
        if pd.notna(max_elevation):
            chart.x_axis.scaling.max = max(5, math.ceil(float(max_elevation)))

    headers = list(output_frame.columns)
    elevation_cols: Dict[str, int] = {}
    snr_cols: Dict[str, int] = {}
    for idx, name in enumerate(headers, start=1):
        if name.endswith("6_2_elevation"):
            elevation_cols[name.split(" ")[1]] = idx
        elif name.endswith("5_10_signal_noise_ratio"):
            snr_cols[name.split(" ")[1]] = idx

    data_start_row = 3
    data_end_row = 2 + output_frame.shape[0]
    for orbit, snr_col in sorted(snr_cols.items(), key=lambda item: int(item[0])):
        elev_col = elevation_cols.get(orbit)
        if elev_col is None:
            continue
        values = Reference(
            source_sheet,
            min_col=snr_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        xvalues = Reference(
            source_sheet,
            min_col=elev_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        series = Series(values, xvalues, title=f"Orbit {orbit}")
        chart.series.append(series)

    chart_sheet.add_chart(chart, "A1")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Process MEOS Extract Excel files and consolidate "
            "UTC time, SNR, azimuth, and elevation."
        )
    )
    parser.add_argument(
        "folder",
        nargs="?",
        default=".",
        help="Folder containing MEOS Extract Excel files.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="meos_processed.xlsx",
        help="Output Excel file path.",
    )
    parser.add_argument(
        "--azimuth-tolerance",
        type=float,
        default=0.1,
        help="Tolerance (degrees) for matching azimuth values between orbits.",
    )
    parser.add_argument(
        "--elevation-tolerance",
        type=float,
        default=0.1,
        help="Tolerance (degrees) for matching elevation values between orbits.",
    )
    args = parser.parse_args()

    folder = Path(args.folder).resolve()
    if not folder.exists() or not folder.is_dir():
        raise SystemExit(f"Folder not found: {folder}")

    output_path = Path(args.output).resolve()
    excel_files = [
        *folder.glob("*.xlsx"),
        *folder.glob("*.xls"),
        *folder.glob("*.xlsm"),
    ]
    excel_files = [path for path in excel_files if path.resolve() != output_path]
    if not excel_files:
        raise SystemExit(f"No Excel files found in {folder}")

    processed_frames: List[Tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]] = []
    for path in sorted(excel_files):
        print(f"Processing Excel file: {path.name}")
        frame = _process_file(path)
        start_time = _extract_start_time(path)
        group_time = start_time.floor("h")
        processed_frames.append((frame, start_time, group_time))

    existing_sheets: Dict[str, pd.DataFrame] = {}
    if output_path.exists():
        existing_sheets = pd.read_excel(output_path, sheet_name=None, header=1)

    origin = min(group_time for _, _, group_time in processed_frames)
    nine_days = pd.Timedelta(days=9)
    grouped_frames: Dict[int, List[Tuple[pd.DataFrame, pd.Timestamp]]] = {}
    for frame, start_time, group_time in processed_frames:
        remainder = (group_time - origin) % nine_days
        key = int(remainder.total_seconds())
        grouped_frames.setdefault(key, []).append((frame, start_time))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for idx, key in enumerate(sorted(grouped_frames)):
            sheet_name = f"Group {idx + 1}"
            group_items = sorted(grouped_frames[key], key=lambda item: item[1])
            output_frame = existing_sheets.get(sheet_name)
            if output_frame is not None:
                output_frame = output_frame.reset_index(drop=True)
            for frame, _ in group_items:
                orbit = int(frame["orbit"].iloc[0])
                columns = [
                    "time_iso_utc",
                    "6_1_azimuth",
                    "6_2_elevation",
                    "5_10_signal_noise_ratio",
                ]
                block = frame[columns].copy().reset_index(drop=True)
                block["time_iso_utc"] = block["time_iso_utc"].dt.tz_localize(None)
                block = block.rename(
                    columns={
                        "time_iso_utc": f"Orbit {orbit} time_iso_utc",
                        "6_1_azimuth": f"Orbit {orbit} 6_1_azimuth",
                        "6_2_elevation": f"Orbit {orbit} 6_2_elevation",
                        "5_10_signal_noise_ratio": (
                            f"Orbit {orbit} 5_10_signal_noise_ratio"
                        ),
                    }
                )
                if output_frame is None:
                    output_frame = block
                else:
                    output_frame = pd.concat(
                        [output_frame, block], axis=1, ignore_index=False
                    )
            if output_frame is None:
                output_frame = pd.DataFrame()
            output_frame.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=1,
            )
            worksheet = writer.sheets[sheet_name]
            if output_frame.shape[1] > 0:
                worksheet.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=output_frame.shape[1],
                )
                title_cell = worksheet.cell(row=1, column=1)
                title_cell.value = sheet_name
                title_cell.alignment = Alignment(horizontal="center")

            _add_group_chart_sheet(writer.book, worksheet, sheet_name, output_frame)

            frames_only = [frame for frame, _ in group_items]
            correlation_frame = _build_correlation_table(
                frames_only,
                args.azimuth_tolerance,
                args.elevation_tolerance,
            )
            if not correlation_frame.empty:
                data_end_row = 2 + output_frame.shape[0]
                title_row = data_end_row + 2
                worksheet.cell(
                    row=title_row,
                    column=1,
                    value="SNR Correlation (Azimuth/Elevation matched)",
                )
                header_row = title_row + 1
                for row_idx, row in enumerate(
                    dataframe_to_rows(correlation_frame, index=False, header=True),
                    start=header_row,
                ):
                    for col_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

                chart = ScatterChart()
                chart.title = "SNR by Orbit (Matched Az/El)"
                chart.x_axis.title = "Azimuth (binned)"
                chart.y_axis.title = "SNR"
                data_start_row = header_row + 1
                data_end_row = data_start_row + correlation_frame.shape[0] - 1
                azimuth_col = 1
                for col in range(3, correlation_frame.shape[1] + 1):
                    values = Reference(
                        worksheet,
                        min_col=col,
                        min_row=data_start_row,
                        max_row=data_end_row,
                    )
                    xvalues = Reference(
                        worksheet,
                        min_col=azimuth_col,
                        min_row=data_start_row,
                        max_row=data_end_row,
                    )
                    series = Series(values, xvalues, title_from_data=False)
                    label_value = worksheet.cell(row=header_row, column=col).value
                    if label_value is not None:
                        series.title = SeriesLabel(v=str(label_value))
                    chart.series.append(series)
                chart_anchor = f"A{data_end_row + 2}"
                worksheet.add_chart(chart, chart_anchor)

    total_rows = sum(len(frame) for frame, _, _ in processed_frames)
    print(f"Saved {total_rows} rows to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
